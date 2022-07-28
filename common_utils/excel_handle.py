# -*- coding: utf-8 -*-

# @version: v1.0
# @author : Hide
# @Project : basic
# @File : excel_handle.py
# @Software: PyCharm
# @time: 2022/7/19 14:16
# @description :
from openpyxl import load_workbook
from common_utils.global_vars import CaseEnum


class ExcelHandle:

    def __init__(self, filename):
        self.wb = load_workbook(filename)
        self.sheet_names = self.wb.sheetnames

    def read_excel(self, sheet=None):
        if sheet is None:
            sheet_names = self.sheet_names
        else:
            sheet_names = sheet
        case_data = []
        for sheet_name in sheet_names:
            max_row = self.wb[sheet_name].max_row
            for row in [row for row in range(2, max_row + 1) if
                      self.wb[sheet_name].cell(row, CaseEnum.API_EXEC.value).value == "是"]:
                _dict = {}
                _dict["id"] = self.wb[sheet_name].cell(row, column=CaseEnum.CASE_ID.value).value
                _dict["feature"] = self.wb[sheet_name].cell(row, column=CaseEnum.CASE_FEATURE.value).value
                _dict["title"] = self.wb[sheet_name].cell(row, column=CaseEnum.CASE_TITLE.value).value
                _dict["url"] = self.wb[sheet_name].cell(row, column=CaseEnum.API_PATH.value).value
                _dict["header"] = self.wb[sheet_name].cell(row, column=CaseEnum.API_HEADER.value).value
                _dict["method"] = self.wb[sheet_name].cell(row, column=CaseEnum.API_METHOD.value).value
                _dict["pk"] = self.wb[sheet_name].cell(row, column=CaseEnum.API_PK.value).value
                _dict["data"] = self.wb[sheet_name].cell(row, column=CaseEnum.API_DATA.value).value
                _dict["file"] = self.wb[sheet_name].cell(row, column=CaseEnum.API_FILE.value).value
                _dict["extract"] = self.wb[sheet_name].cell(row, column=CaseEnum.API_EXTRACT.value).value
                _dict["validate"] = self.wb[sheet_name].cell(row, column=CaseEnum.API_EXPECTED.value).value

                case_data.append(_dict)
        return case_data


if __name__ == '__main__':
    excel_result = ExcelHandle("../data/test_demo1.xlsx").read_excel()
    print(excel_result)